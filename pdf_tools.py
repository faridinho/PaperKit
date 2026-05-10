from __future__ import annotations

import hashlib
import re
import shutil
import subprocess
import sys
from datetime import datetime
from pathlib import Path
from threading import Event
from typing import Callable, Optional, Any

import fitz  # PyMuPDF
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

LogFunc = Callable[[str], None]
ProgressFunc = Callable[[int, int], None]

GHOSTSCRIPT_DOWNLOAD_URL = "https://ghostscript.com/releases/gsdnld.html"

PDF_SETTINGS = {
    "screen": "/screen",      # lowest quality, smallest file
    "ebook": "/ebook",        # good balance
    "printer": "/printer",    # higher quality
    "prepress": "/prepress",  # highest quality, larger file
}

METADATA_FIELDS = [
    "ID",
    "Filename",
    "Title",
    "Author",
    "Year",
    "DOI",
    "Page count",
    "File size MB",
    "Creation date",
    "Modification date",
    "Subject",
    "Keywords",
    "Creator",
    "Producer",
]

DEFAULT_METADATA_FIELDS = ["ID", "Title", "Author", "Year"]


def default_log(message: str) -> None:
    print(message)


def clean_filename(text: str, max_length: int = 150, lowercase_rest: bool = False) -> str:
    text = re.sub(r'[\\/*?:"<>|]', "", text)
    text = re.sub(r"\s+", " ", text).strip()
    if not text:
        text = "Untitled"
    if lowercase_rest:
        text = text.capitalize()
    return text[:max_length].strip()


def make_unique_path(path: Path) -> Path:
    if not path.exists():
        return path
    parent = path.parent
    base = path.stem
    suffix = path.suffix
    i = 2
    while True:
        candidate = parent / f"{base} ({i}){suffix}"
        if not candidate.exists():
            return candidate
        i += 1


def list_pdfs(folder: Path) -> list[Path]:
    return sorted(Path(folder).glob("*.pdf"), key=lambda p: p.name.lower())


def extract_title(pdf_path: Path, lowercase_rest: bool = False) -> str:
    doc = fitz.open(pdf_path)
    try:
        metadata_title = (doc.metadata or {}).get("title", "") or ""
        if metadata_title and len(metadata_title.strip()) > 10:
            return clean_filename(metadata_title, lowercase_rest=lowercase_rest)

        if len(doc) == 0:
            return "Untitled"

        first_page_text = doc[0].get_text("text")
        lines = [line.strip() for line in first_page_text.splitlines() if len(line.strip()) > 10]

        skip_keywords = [
            "research article", "accepted", "submitted", "doi", "issn", "journal",
            "volume", "issue", "published online", "crossmark", "copyright",
            "special collection", "proceedings", "abstract", "keywords",
        ]

        filtered = [
            line for line in lines
            if not any(keyword in line.lower() for keyword in skip_keywords)
        ]

        candidate = max(filtered, key=lambda line: len(line.split()), default="Untitled")
        return clean_filename(candidate, lowercase_rest=lowercase_rest)
    finally:
        doc.close()


def build_rename_plan(
    input_dir: Path,
    output_dir: Path,
    start_number: int = 1,
    add_numbering: bool = True,
    in_place: bool = False,
    lowercase_rest: bool = False,
    log: LogFunc = default_log,
    progress: Optional[ProgressFunc] = None,
) -> list[dict[str, Any]]:
    input_dir = Path(input_dir)
    output_dir = Path(output_dir)
    if not input_dir.exists():
        raise FileNotFoundError(f"Input folder does not exist: {input_dir}")

    pdf_files = list_pdfs(input_dir)
    plan: list[dict[str, Any]] = []
    seen_targets: set[str] = set()
    total = len(pdf_files)

    if total == 0:
        log("No PDF files found for renaming preview.")
        return plan

    destination_dir = input_dir if in_place else output_dir / "renamed"

    for index, pdf_path in enumerate(pdf_files, start=1):
        try:
            title = extract_title(pdf_path, lowercase_rest=lowercase_rest)
            if add_numbering:
                new_name = f"{start_number + index - 1}-{title}.pdf"
            else:
                new_name = f"{title}.pdf"

            target_path = destination_dir / new_name
            status = "Ready"

            normalized_target = str(target_path).lower()
            if normalized_target in seen_targets:
                status = "Duplicate target name; will make unique when applied"
            seen_targets.add(normalized_target)

            if target_path.exists() and target_path.resolve() != pdf_path.resolve():
                status = "Target exists; will make unique when applied"

            plan.append({
                "source_path": str(pdf_path),
                "target_dir": str(destination_dir),
                "old_filename": pdf_path.name,
                "title": title,
                "new_filename": new_name,
                "status": status,
            })
            log(f"Preview: {pdf_path.name} -> {new_name}")
        except Exception as exc:
            plan.append({
                "source_path": str(pdf_path),
                "target_dir": str(destination_dir),
                "old_filename": pdf_path.name,
                "title": "",
                "new_filename": "",
                "status": f"ERROR: {exc}",
            })
            log(f"FAILED preview for {pdf_path.name}: {exc}")
        finally:
            if progress:
                progress(index, total)

    return plan


def apply_rename_plan(
    plan: list[dict[str, Any]],
    in_place: bool = False,
    log: LogFunc = default_log,
    progress: Optional[ProgressFunc] = None,
) -> list[Path]:
    if not plan:
        log("No rename preview plan to apply.")
        return []

    renamed: list[Path] = []
    total = len(plan)

    for index, item in enumerate(plan, start=1):
        try:
            if str(item.get("status", "")).startswith("ERROR"):
                log(f"Skipped {item.get('old_filename')}: preview had an error.")
                continue

            source = Path(item["source_path"])
            target_dir = Path(item["target_dir"])
            target_dir.mkdir(parents=True, exist_ok=True)
            target = make_unique_path(target_dir / item["new_filename"])

            if in_place:
                if source.resolve() == target.resolve():
                    log(f"Skipped already named: {source.name}")
                    renamed.append(source)
                else:
                    source.rename(target)
                    renamed.append(target)
                    log(f"Renamed: {source.name} -> {target.name}")
            else:
                shutil.copy2(source, target)
                renamed.append(target)
                log(f"Copied + renamed: {source.name} -> {target.name}")
        except Exception as exc:
            log(f"FAILED to apply rename for {item.get('old_filename')}: {exc}")
        finally:
            if progress:
                progress(index, total)

    return renamed


def file_sha256(path: Path, chunk_size: int = 1024 * 1024) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as file:
        for chunk in iter(lambda: file.read(chunk_size), b""):
            h.update(chunk)
    return h.hexdigest()


def find_duplicate_pdfs_by_hash(
    input_dir: Path,
    log: LogFunc = default_log,
    progress: Optional[ProgressFunc] = None,
) -> list[dict[str, Any]]:
    input_dir = Path(input_dir)
    pdf_files = list_pdfs(input_dir)
    total = len(pdf_files)
    by_hash: dict[str, list[Path]] = {}

    for index, pdf_path in enumerate(pdf_files, start=1):
        try:
            digest = file_sha256(pdf_path)
            by_hash.setdefault(digest, []).append(pdf_path)
            log(f"Scanned for duplicates: {pdf_path.name}")
        except Exception as exc:
            log(f"FAILED duplicate scan for {pdf_path.name}: {exc}")
        finally:
            if progress:
                progress(index, total)

    groups = []
    group_num = 1
    for digest, paths in by_hash.items():
        if len(paths) > 1:
            for path in paths:
                groups.append({
                    "Group": group_num,
                    "Filename": path.name,
                    "File path": str(path),
                    "File size MB": round(path.stat().st_size / 1024 / 1024, 3),
                    "SHA256": digest,
                    "Suggested action": "Review duplicate",
                })
            group_num += 1

    if not groups:
        log("No exact duplicate PDFs found by file hash.")
    else:
        log(f"Found {group_num - 1} duplicate group(s).")

    return groups


def export_duplicate_report(rows: list[dict[str, Any]], output_file: Path, log: LogFunc = default_log) -> Path:
    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Duplicate PDFs"
    headers = ["Group", "Filename", "File path", "File size MB", "SHA256", "Suggested action"]
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    for col_idx, header in enumerate(headers, start=1):
        width = 18
        if header in {"Filename", "File path", "SHA256"}:
            width = 60
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    wb.save(output_file)
    log(f"Duplicate report saved: {output_file}")
    return output_file


def find_ghostscript(gs_executable: Optional[str] = None) -> str:
    candidates: list[str] = []
    if gs_executable:
        candidates.append(gs_executable)
    if sys.platform.startswith("win"):
        candidates.extend(["gswin64c", "gswin32c", "gs"])
    else:
        candidates.append("gs")

    for candidate in candidates:
        resolved = shutil.which(candidate)
        if resolved:
            return resolved
        path_candidate = Path(candidate)
        if path_candidate.exists():
            return str(path_candidate)

    raise RuntimeError(
        "Ghostscript was not found.\n\n"
        "Compression requires Ghostscript.\n"
        f"Download it here: {GHOSTSCRIPT_DOWNLOAD_URL}\n\n"
        "After installing, either leave the Ghostscript field empty if gswin64c is in PATH, "
        "or paste the full path to gswin64c.exe."
    )


def compress_pdf(
    pdf_path: Path,
    output_path: Path,
    quality: str = "ebook",
    gs_executable: Optional[str] = None,
    log: LogFunc = default_log,
    cancel_event: Optional[Event] = None,
) -> None:
    if quality not in PDF_SETTINGS:
        raise ValueError(f"Invalid quality: {quality}")

    if cancel_event and cancel_event.is_set():
        raise RuntimeError("Compression cancelled.")

    gs = find_ghostscript(gs_executable)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    command = [
        gs,
        "-sDEVICE=pdfwrite",
        "-dCompatibilityLevel=1.4",
        f"-dPDFSETTINGS={PDF_SETTINGS[quality]}",
        "-dNOPAUSE",
        "-dQUIET",
        "-dBATCH",
        f"-sOutputFile={str(output_path)}",
        str(pdf_path),
    ]

    process = subprocess.Popen(
        command,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
        encoding="utf-8",
        errors="replace",
    )

    while process.poll() is None:
        if cancel_event and cancel_event.is_set():
            process.terminate()
            try:
                process.wait(timeout=5)
            except subprocess.TimeoutExpired:
                process.kill()
            if output_path.exists():
                try:
                    output_path.unlink()
                except OSError:
                    pass
            raise RuntimeError("Compression cancelled by user.")

    stdout, stderr = process.communicate()
    if process.returncode != 0:
        raise RuntimeError(f"Ghostscript failed for {pdf_path.name}.\n\n{stderr.strip() or stdout.strip()}")

    original_mb = pdf_path.stat().st_size / 1024 / 1024
    compressed_mb = output_path.stat().st_size / 1024 / 1024
    reduction = 100 * (1 - compressed_mb / original_mb) if original_mb else 0
    log(f"Compressed: {pdf_path.name}: {original_mb:.2f} MB -> {compressed_mb:.2f} MB ({reduction:.1f}% smaller)")


def compress_pdfs_in_folder(
    input_dir: Path,
    output_dir: Path,
    quality: str = "ebook",
    gs_executable: Optional[str] = None,
    log: LogFunc = default_log,
    progress: Optional[ProgressFunc] = None,
    cancel_event: Optional[Event] = None,
) -> list[Path]:
    input_dir = Path(input_dir)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    pdf_files = list_pdfs(input_dir)
    compressed: list[Path] = []
    total = len(pdf_files)

    if total == 0:
        log("No PDF files found for compression.")
        return compressed

    for index, pdf_path in enumerate(pdf_files, start=1):
        if cancel_event and cancel_event.is_set():
            log("Compression cancelled before next file.")
            break
        try:
            out_path = make_unique_path(output_dir / pdf_path.name)
            compress_pdf(pdf_path, out_path, quality, gs_executable, log, cancel_event)
            compressed.append(out_path)
        except Exception as exc:
            if cancel_event and cancel_event.is_set():
                log(str(exc))
                break
            log(f"FAILED to compress {pdf_path.name}: {exc}")
        finally:
            if progress:
                progress(index, total)
    return compressed


def parse_id_from_filename(pdf_path: Path) -> str:
    match = re.match(r"^(\d+)-", pdf_path.name)
    return match.group(1) if match else ""


def parse_year_from_pdf_date(date_string: str) -> str:
    if not date_string:
        return ""
    match = re.search(r"(19|20)\d{2}", date_string)
    return match.group(0) if match else ""


def extract_year_from_text(text: str) -> str:
    years = re.findall(r"\b(19\d{2}|20\d{2})\b", text)
    current_year = datetime.now().year + 1
    valid_years = [int(y) for y in years if 1900 <= int(y) <= current_year]
    return str(min(valid_years)) if valid_years else ""


def extract_doi_from_text(text: str) -> str:
    # DOI pattern recommended style: 10.<prefix>/<suffix>; stops before whitespace and common brackets.
    match = re.search(r"\b10\.\d{4,9}/[^\s\)\]\}\>,;]+", text, flags=re.IGNORECASE)
    if not match:
        return ""
    doi = match.group(0).rstrip(".:")
    return doi


def normalize_pdf_date(date_string: str) -> str:
    if not date_string:
        return ""
    m = re.search(r"D?[:]?(\d{4})(\d{2})?(\d{2})?", date_string)
    if not m:
        return date_string
    year, month, day = m.group(1), m.group(2), m.group(3)
    if month and day:
        return f"{year}-{month}-{day}"
    if month:
        return f"{year}-{month}"
    return year


def extract_metadata_for_pdf(pdf_path: Path) -> dict[str, Any]:
    doc = fitz.open(pdf_path)
    try:
        meta = doc.metadata or {}
        first_page_text = doc[0].get_text("text") if len(doc) > 0 else ""
        all_front_text = first_page_text

        title = (meta.get("title", "") or "").strip() or extract_title(pdf_path)
        author = (meta.get("author", "") or "").strip()
        creation = meta.get("creationDate", "") or ""
        modification = meta.get("modDate", "") or ""
        year = parse_year_from_pdf_date(creation) or parse_year_from_pdf_date(modification) or extract_year_from_text(first_page_text) or extract_year_from_text(pdf_path.name)
        doi = extract_doi_from_text(all_front_text) or extract_doi_from_text(" ".join(str(v) for v in meta.values()))

        return {
            "ID": parse_id_from_filename(pdf_path),
            "Filename": pdf_path.name,
            "Title": title,
            "Author": author,
            "Year": year,
            "DOI": doi,
            "Page count": len(doc),
            "File size MB": round(pdf_path.stat().st_size / 1024 / 1024, 3),
            "Creation date": normalize_pdf_date(creation),
            "Modification date": normalize_pdf_date(modification),
            "Subject": (meta.get("subject", "") or "").strip(),
            "Keywords": (meta.get("keywords", "") or "").strip(),
            "Creator": (meta.get("creator", "") or "").strip(),
            "Producer": (meta.get("producer", "") or "").strip(),
        }
    finally:
        doc.close()


def export_metadata_to_excel(
    input_dir: Path,
    output_file: Path,
    fields: Optional[list[str]] = None,
    log: LogFunc = default_log,
    progress: Optional[ProgressFunc] = None,
) -> Path:
    input_dir = Path(input_dir)
    fields = fields or DEFAULT_METADATA_FIELDS
    invalid = [field for field in fields if field not in METADATA_FIELDS]
    if invalid:
        raise ValueError(f"Invalid metadata fields: {invalid}")

    pdf_files = list_pdfs(input_dir)
    output_file.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "PDF Metadata"
    ws.append(fields)

    total = len(pdf_files)
    if total == 0:
        log("No PDF files found for metadata extraction.")
    for index, pdf_path in enumerate(pdf_files, start=1):
        try:
            data = extract_metadata_for_pdf(pdf_path)
            ws.append([data.get(field, "") for field in fields])
            log(f"Metadata extracted: {pdf_path.name}")
        except Exception as exc:
            log(f"FAILED metadata extraction for {pdf_path.name}: {exc}")
            fallback = {field: "" for field in fields}
            if "Filename" in fallback:
                fallback["Filename"] = pdf_path.name
            if "Title" in fallback:
                fallback["Title"] = pdf_path.stem
            ws.append([fallback.get(field, "") for field in fields])
        finally:
            if progress:
                progress(index, total)

    for col_idx, field in enumerate(fields, start=1):
        width = 16
        if field in {"Title", "Filename", "Author", "DOI", "Keywords"}:
            width = 45 if field != "Title" else 70
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    wb.save(output_file)
    log(f"Excel metadata file created: {output_file}")
    return output_file



# -----------------------------
# Citation export helpers
# -----------------------------

def _safe_bibtex_key(text: str) -> str:
    text = re.sub(r"[^A-Za-z0-9]+", "_", text).strip("_")
    return text or "paper"


def split_authors(author_text: str) -> list[str]:
    """Best-effort author splitting for PDF metadata strings."""
    author_text = (author_text or "").strip()
    if not author_text:
        return []

    # PDF author metadata may use semicolons, commas, pipes, or "and".
    # Keep this conservative to avoid over-splitting names like "Smith, John".
    if ";" in author_text:
        parts = author_text.split(";")
    elif " | " in author_text:
        parts = author_text.split(" | ")
    elif " and " in author_text.lower():
        parts = re.split(r"\s+and\s+", author_text, flags=re.IGNORECASE)
    else:
        parts = [author_text]

    return [p.strip() for p in parts if p.strip()]


def _ris_escape(value: Any) -> str:
    return str(value or "").replace("\n", " ").replace("\r", " ").strip()


def metadata_to_ris_entry(data: dict[str, Any]) -> str:
    lines = ["TY  - JOUR"]

    title = _ris_escape(data.get("Title"))
    year = _ris_escape(data.get("Year"))
    doi = _ris_escape(data.get("DOI"))
    filename = _ris_escape(data.get("Filename"))
    keywords = _ris_escape(data.get("Keywords"))

    if title:
        lines.append(f"TI  - {title}")

    for author in split_authors(_ris_escape(data.get("Author"))):
        lines.append(f"AU  - {author}")

    if year:
        lines.append(f"PY  - {year}")
    if doi:
        lines.append(f"DO  - {doi}")
    if keywords:
        for keyword in re.split(r"[,;]", keywords):
            keyword = keyword.strip()
            if keyword:
                lines.append(f"KW  - {keyword}")
    if filename:
        lines.append(f"N1  - Source PDF: {filename}")

    lines.append("ER  -")
    return "\n".join(lines)


def _bibtex_escape(value: Any) -> str:
    text = str(value or "").replace("\n", " ").replace("\r", " ").strip()
    replacements = {
        "\\": r"\\",
        "{": r"\{",
        "}": r"\}",
        "&": r"\&",
        "%": r"\%",
        "$": r"\$",
        "#": r"\#",
        "_": r"\_",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    return text


def metadata_to_bibtex_entry(data: dict[str, Any], index: int) -> str:
    title_raw = str(data.get("Title") or "paper")
    year_raw = str(data.get("Year") or "n.d.")
    first_author = split_authors(str(data.get("Author") or ""))
    author_key = first_author[0].split()[-1] if first_author else "paperkit"
    key = _safe_bibtex_key(f"{author_key}_{year_raw}_{index}")

    fields = []
    title = _bibtex_escape(data.get("Title"))
    author = _bibtex_escape(data.get("Author"))
    year = _bibtex_escape(data.get("Year"))
    doi = _bibtex_escape(data.get("DOI"))
    keywords = _bibtex_escape(data.get("Keywords"))
    filename = _bibtex_escape(data.get("Filename"))

    if title:
        fields.append(f"  title = {{{title}}}")
    if author:
        # BibTeX expects authors joined by "and". Keep metadata as-is if uncertain.
        authors = split_authors(str(data.get("Author") or ""))
        author_value = " and ".join(authors) if authors else str(data.get("Author") or "")
        fields.append(f"  author = {{{_bibtex_escape(author_value)}}}")
    if year:
        fields.append(f"  year = {{{year}}}")
    if doi:
        fields.append(f"  doi = {{{doi}}}")
    if keywords:
        fields.append(f"  keywords = {{{keywords}}}")
    if filename:
        fields.append(f"  note = {{Source PDF: {filename}}}")

    if not fields:
        fields.append(f"  title = {{{_bibtex_escape(title_raw)}}}")

    return "@article{" + key + ",\n" + ",\n".join(fields) + "\n}"


def export_citations(
    input_dir: Path,
    output_dir: Path,
    export_ris: bool = True,
    export_bibtex: bool = True,
    log: LogFunc = default_log,
    progress: Optional[ProgressFunc] = None,
) -> list[Path]:
    """
    Automatically extracts available PDF metadata and exports citation files.
    Output quality depends on the metadata embedded in the PDFs and detectable DOI/title text.
    """
    input_dir = Path(input_dir)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    pdf_files = list_pdfs(input_dir)
    total = len(pdf_files)
    records: list[dict[str, Any]] = []

    if total == 0:
        log("No PDF files found for citation export.")
        return []

    for index, pdf_path in enumerate(pdf_files, start=1):
        try:
            data = extract_metadata_for_pdf(pdf_path)
            records.append(data)
            missing = []
            for required in ["Title", "Author", "Year", "DOI"]:
                if not data.get(required):
                    missing.append(required)
            if missing:
                log(f"Citation metadata extracted with missing fields for {pdf_path.name}: {', '.join(missing)}")
            else:
                log(f"Citation metadata extracted: {pdf_path.name}")
        except Exception as exc:
            log(f"FAILED citation extraction for {pdf_path.name}: {exc}")
        finally:
            if progress:
                progress(index, total)

    outputs: list[Path] = []

    if export_ris:
        ris_path = output_dir / "paperkit_citations.ris"
        ris_text = "\n\n".join(metadata_to_ris_entry(record) for record in records)
        ris_path.write_text(ris_text + ("\n" if ris_text else ""), encoding="utf-8")
        outputs.append(ris_path)
        log(f"RIS citation file created: {ris_path}")

    if export_bibtex:
        bib_path = output_dir / "paperkit_citations.bib"
        bib_text = "\n\n".join(metadata_to_bibtex_entry(record, i) for i, record in enumerate(records, start=1))
        bib_path.write_text(bib_text + ("\n" if bib_text else ""), encoding="utf-8")
        outputs.append(bib_path)
        log(f"BibTeX citation file created: {bib_path}")

    log("Citation export finished. Review imported records in your reference manager because PDF metadata can be incomplete.")
    return outputs

def scan_embedded_images(pdf_path: Path, log: LogFunc = default_log) -> list[dict[str, Any]]:
    pdf_path = Path(pdf_path)
    if not pdf_path.exists():
        raise FileNotFoundError(f"PDF file not found: {pdf_path}")
    if pdf_path.suffix.lower() != ".pdf":
        raise ValueError("Please choose a PDF file.")

    doc = fitz.open(pdf_path)
    images: list[dict[str, Any]] = []
    seen_xrefs: set[int] = set()
    try:
        for page_index in range(len(doc)):
            page = doc[page_index]
            page_image_index = 0
            for info in page.get_images(full=True):
                xref = info[0]
                if xref in seen_xrefs:
                    continue
                seen_xrefs.add(xref)
                try:
                    extracted = doc.extract_image(xref)
                    page_image_index += 1
                    images.append({
                        "xref": xref,
                        "page": page_index + 1,
                        "image_number": page_image_index,
                        "width": extracted.get("width", ""),
                        "height": extracted.get("height", ""),
                        "format": extracted.get("ext", "png"),
                        "size_bytes": len(extracted.get("image", b"")),
                    })
                except Exception as exc:
                    log(f"FAILED to scan image xref {xref} on page {page_index + 1}: {exc}")
        log(f"Found {len(images)} extractable embedded image(s).")
        return images
    finally:
        doc.close()


def get_image_bytes(pdf_path: Path, xref: int) -> tuple[bytes, str]:
    doc = fitz.open(pdf_path)
    try:
        extracted = doc.extract_image(xref)
        return extracted["image"], extracted.get("ext", "png")
    finally:
        doc.close()


def extract_selected_images_from_pdf(
    pdf_path: Path,
    output_dir: Path,
    xrefs: list[int],
    log: LogFunc = default_log,
    progress: Optional[ProgressFunc] = None,
) -> Path:
    pdf_path = Path(pdf_path)
    output_dir = Path(output_dir)
    if not xrefs:
        raise ValueError("No images selected.")
    safe_pdf_name = clean_filename(pdf_path.stem, max_length=80)
    image_folder = output_dir / f"{safe_pdf_name}_images"
    image_folder.mkdir(parents=True, exist_ok=True)

    doc = fitz.open(pdf_path)
    total = len(xrefs)
    try:
        xref_set = set(xrefs)
        page_lookup: dict[int, int] = {}
        for page_index in range(len(doc)):
            for info in doc[page_index].get_images(full=True):
                xref = info[0]
                if xref in xref_set and xref not in page_lookup:
                    page_lookup[xref] = page_index + 1

        for index, xref in enumerate(xrefs, start=1):
            try:
                extracted = doc.extract_image(xref)
                image_bytes = extracted["image"]
                ext = extracted.get("ext", "png")
                width = extracted.get("width", "")
                height = extracted.get("height", "")
                page_num = page_lookup.get(xref, 0)
                filename = f"page_{page_num:03d}_image_{index:03d}_xref_{xref}_{width}x{height}.{ext}"
                path = make_unique_path(image_folder / filename)
                with open(path, "wb") as file:
                    file.write(image_bytes)
                log(f"Extracted: {path.name}")
            except Exception as exc:
                log(f"FAILED to extract image xref {xref}: {exc}")
            finally:
                if progress:
                    progress(index, total)
        log(f"Selected images saved in: {image_folder}")
        return image_folder
    finally:
        doc.close()


def extract_images_from_pdf(pdf_path: Path, output_dir: Path, log: LogFunc = default_log) -> Path:
    images = scan_embedded_images(pdf_path, log=log)
    return extract_selected_images_from_pdf(pdf_path, output_dir, [int(i["xref"]) for i in images], log=log)

# -----------------------------
# Operation reports
# -----------------------------

def export_operation_report(
    output_file: Path,
    title: str,
    summary: dict[str, Any],
    details: Optional[list[dict[str, Any]]] = None,
    log: LogFunc = default_log,
) -> Path:
    """Create a simple Excel report for a PaperKit operation."""
    output_file = Path(output_file)
    output_file.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.append([title])
    ws.append([])
    ws.append(["Item", "Value"])
    for key, value in summary.items():
        ws.append([key, value])
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 80

    if details:
        ws2 = wb.create_sheet("Details")
        headers = list(details[0].keys())
        ws2.append(headers)
        for row in details:
            ws2.append([row.get(h, "") for h in headers])
        for idx, header in enumerate(headers, start=1):
            width = 18
            if any(word in header.lower() for word in ["file", "path", "title", "error", "status"]):
                width = 50
            ws2.column_dimensions[get_column_letter(idx)].width = width

    wb.save(output_file)
    log(f"Operation report saved: {output_file}")
    return output_file
